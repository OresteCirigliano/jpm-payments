import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from payments.iban_validator import validate_iban

PAYABLE_EXCLUDE = [0, 10]

COUNTRY_SPECIAL = {
    'BE': {'sodexo_payable': 5,    'sodexo_swift': None,     'sodexo_label': 'Payquiker'},
    'NL': {'sodexo_payable': 5,    'sodexo_swift': None,     'sodexo_label': 'Payquiker'},
    'PL': {'sodexo_payable': 8,    'sodexo_swift': 'SODEXO', 'sodexo_label': 'Sodexo'},
}

def validate(df_emea, country_code, emea_filter_code, generated_ids, generated_totals, sodexo_exclude=False):
    df = df_emea[df_emea['Country'].str.strip().str.upper() == emea_filter_code.upper()].copy()
    df['effective_id'] = df['CustomerID'].astype(str).str.strip()
    df['PayableTy']    = pd.to_numeric(df['PayableTy'], errors='coerce')
    df['Field11']      = pd.to_numeric(df['Field11'],   errors='coerce')
    df['Amount']       = pd.to_numeric(df['Amount'],    errors='coerce').fillna(0)

    emea_totals = df.groupby('effective_id')['Amount'].sum().round(2)
    names       = df.groupby('effective_id')['DepositName'].first()
    ibans       = df.groupby('effective_id')['IBAN'].first()
    bill_counts = df.groupby('effective_id').size()

    special      = COUNTRY_SPECIAL.get(country_code, {})
    sodexo_pay   = special.get('sodexo_payable', None)
    sodexo_swft  = special.get('sodexo_swift', None)
    sodexo_label = special.get('sodexo_label', 'Sodexo')

    exclusions_normal = []
    anomalies         = []
    all_emea_ids      = set(df['effective_id'].unique())

    for cid in all_emea_ids:
        rows       = df[df['effective_id'] == cid]
        total_emea = round(emea_totals.get(cid, 0), 2)

        payable_vals      = rows['PayableTy'].dropna().unique().tolist()
        field11_vals      = rows['Field11'].dropna().unique().tolist()
        has_payable_0_10  = any(v in PAYABLE_EXCLUDE for v in payable_vals)
        has_field11_block = any(v not in [3] for v in field11_vals if pd.notna(v))

        has_sodexo = False
        if sodexo_exclude and 5 in payable_vals:
            has_sodexo = True
        if sodexo_pay is not None and sodexo_pay in payable_vals:
            has_sodexo = True
        if sodexo_swft is not None:
            swift_vals = rows['SwiftCode'].astype(str).str.strip().str.upper()
            if swift_vals.eq(sodexo_swft.upper()).any():
                has_sodexo = True

        iban = str(rows['IBAN'].iloc[0]).strip()
        acct = str(rows['DepositAccountNumber'].iloc[0]).strip()
        has_bank = (
            iban.upper() not in ('NULL', 'NAN', '') or
            (acct.upper() not in ('NULL', 'NAN', '') and acct.strip('0') != '')
        )
        iban_missing = iban.upper() in ('NULL', 'NAN', '')

        if cid in generated_ids:
            total_gen = round(generated_totals.get(cid, 0), 2)
            diff      = round(total_gen - total_emea, 2)
            if abs(diff) > 0.01:
                anomalies.append({
                    'CustomerID':       cid,
                    'Type':             'Amount discrepancy',
                    'EMEA Amount':      total_emea,
                    'Generated Amount': total_gen,
                    'Difference':       diff,
                    'Detail':           f'Expected {total_emea}, generated {total_gen}',
                })
        else:
            name = str(names.get(cid, ''))
            # Priority: negative → zero → hold → field11 → payquiker/sodexo → bank details
            if total_emea < 0:
                exclusions_normal.append({'CustomerID': cid, 'Name': name, 'Reason': 'Negative amount',                        'EMEA Amount': total_emea, 'Paid': 'No - Negative'})
            elif total_emea == 0:
                exclusions_normal.append({'CustomerID': cid, 'Name': name, 'Reason': 'Zero amount',                            'EMEA Amount': total_emea, 'Paid': 'No - Zero'})
            elif has_payable_0_10:
                exclusions_normal.append({'CustomerID': cid, 'Name': name, 'Reason': 'Payable excluded (0 or 10 = hold)',      'EMEA Amount': total_emea, 'Paid': 'No - Hold'})
            elif has_field11_block:
                exclusions_normal.append({'CustomerID': cid, 'Name': name, 'Reason': f'Field11 blocked (value: {field11_vals})','EMEA Amount': total_emea, 'Paid': 'No - Hold'})
            elif has_sodexo:
                exclusions_normal.append({'CustomerID': cid, 'Name': name, 'Reason': f'{sodexo_label} payment (not JPM)',      'EMEA Amount': total_emea, 'Paid': f'Yes - {sodexo_label}'})
            elif not has_bank or iban_missing:
                exclusions_normal.append({'CustomerID': cid, 'Name': name, 'Reason': 'Missing bank details or empty IBAN',     'EMEA Amount': total_emea, 'Paid': 'No - Bank details'})
            else:
                anomalies.append({
                    'CustomerID':       cid,
                    'Type':             'Excluded without clear reason',
                    'EMEA Amount':      total_emea,
                    'Generated Amount': 0,
                    'Difference':       -total_emea,
                    'Detail':           'Present in EMEA but not in generated file without known reason',
                })

    for cid in set(generated_ids) - all_emea_ids:
        anomalies.append({
            'CustomerID':       cid,
            'Type':             'ID not found in EMEA',
            'EMEA Amount':      0,
            'Generated Amount': generated_totals.get(cid, 0),
            'Difference':       generated_totals.get(cid, 0),
            'Detail':           'CustomerID in generated file but not found in EMEA',
        })

    status = 'green' if len(anomalies) == 0 else 'red'

    total_emea_all = round(emea_totals.sum(), 2)
    total_gen_all  = round(sum(generated_totals.values()), 2)

    summary = {
        'status':          status,
        'total_emea':      total_emea_all,
        'total_generated': total_gen_all,
        'diff_total':      round(total_gen_all - total_emea_all, 2),
        'n_emea':          len(all_emea_ids),
        'n_generated':     len(generated_ids),
        'n_exclusions':    len(exclusions_normal),
        'n_anomalies':     len(anomalies),
    }

    payments_list = []
    iban_issues   = 0
    for cid in sorted(generated_ids):
        iban_raw             = str(ibans.get(cid, '')).strip()
        is_valid, emoji, msg = validate_iban(iban_raw, country_code)
        if not is_valid:
            iban_issues += 1
        payments_list.append({
            'CustomerID':  cid,
            'Name':        str(names.get(cid, '')),
            'Amount':      generated_totals.get(cid, 0),
            'Bills':       int(bill_counts.get(cid, 1)),
            'IBAN':        iban_raw,
            'IBAN Status': emoji,
            'IBAN Detail': msg,
        })

    summary['iban_issues'] = iban_issues
    if iban_issues > 0 and status == 'green':
        status = 'yellow'
        summary['status'] = 'yellow'

    buf = _build_report(summary, exclusions_normal, anomalies, payments_list, country_code)
    return status, summary, buf


def _build_report(summary, exclusions_normal, anomalies, payments_list, country_code):
    wb = Workbook()

    GREEN  = 'FF92D050'
    YELLOW = 'FFFFC000'
    RED    = 'FFFF0000'
    HEADER = 'FF4472C4'
    WHITE  = 'FFFFFFFF'
    BLACK  = 'FF000000'

    header_font  = Font(bold=True, color=WHITE)
    header_fill  = PatternFill('solid', fgColor=HEADER)
    center_align = Alignment(horizontal='center')

    def write_header(ws, cols):
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        ws.row_dimensions[1].height = 20

    # --- Foglio 1: Summary ---
    ws1 = wb.active
    ws1.title = '1. Summary'
    # ... (codice del summary invariato) ...
    if summary['status'] == 'green':
        status_text, status_color, font_color = '🟢 OK — No anomalies found', GREEN, WHITE
    elif summary['status'] == 'yellow':
        status_text, status_color, font_color = '🟡 WARNING — IBAN issues detected', YELLOW, BLACK
    else:
        status_text, status_color, font_color = '🔴 WARNING — Anomalies detected!', RED, WHITE

    ws1.merge_cells('A1:C1')
    cell = ws1['A1']
    cell.value, cell.font, cell.fill, cell.alignment = status_text, Font(bold=True, size=14, color=font_color), PatternFill('solid', fgColor=status_color), center_align

    data = [('', '', ''), ('Country', country_code, ''), ('', '', ''),
            ('EMEA Total (all)', summary['total_emea'], ''),
            ('Generated file total', summary['total_generated'], ''),
            ('Difference', summary['diff_total'], '⚠️' if abs(summary['diff_total']) > 0.01 else '✅'),
            ('', '', ''), ('Records in EMEA', summary['n_emea'], ''),
            ('Records in file', summary['n_generated'], ''),
            ('Excluded (known reason)', summary['n_exclusions'], ''),
            ('Anomalies', summary['n_anomalies'], '⚠️' if summary['n_anomalies'] > 0 else '✅'),
            ('IBAN issues', summary.get('iban_issues', 0), '⚠️' if summary.get('iban_issues', 0) > 0 else '✅')]

    for row_idx, (label, value, note) in enumerate(data, 2):
        ws1.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row_idx, column=2, value=value)
        ws1.cell(row=row_idx, column=3, value=note)

    # --- Foglio 2: Payments ---
    ws2 = wb.create_sheet('2. Payments')
    write_header(ws2, ['CustomerID', 'Name', 'Amount', '# Bills', 'IBAN', 'IBAN Status', 'IBAN Detail'])
    for row_idx, p in enumerate(payments_list, 2):
        ws2.cell(row_idx, 1, p['CustomerID']); ws2.cell(row_idx, 2, p['Name'])
        ws2.cell(row_idx, 3, p['Amount']); ws2.cell(row_idx, 4, p['Bills'])
        ws2.cell(row_idx, 5, p['IBAN']).number_format = '@'
        ws2.cell(row_idx, 6, p['IBAN Status']); ws2.cell(row_idx, 7, p['IBAN Detail'])

    # --- Foglio 3: Normal Exclusions (VERSIONE AGGIORNATA) ---
    ws3 = wb.create_sheet('3. Normal exclusions')
    # AGGIUNTE COLONNE 'Name' e 'Paid'
    write_header(ws3, ['CustomerID', 'Name', 'Exclusion reason', 'EMEA Amount', 'Paid'])

    YES_FILL = PatternFill('solid', fgColor='FFE2EFDA') # Verde chiaro
    NO_FILL  = PatternFill('solid', fgColor='FFFFE0E0') # Rosso chiaro

    for row_idx, exc in enumerate(exclusions_normal, 2):
        ws3.cell(row=row_idx, column=1, value=exc['CustomerID'])
        ws3.cell(row=row_idx, column=2, value=exc.get('Name', 'N/A'))
        ws3.cell(row=row_idx, column=3, value=exc['Reason'])
        ws3.cell(row=row_idx, column=4, value=exc['EMEA Amount'])
        
        # Gestione colonna PAID con colori
        paid_val = exc.get('Paid', 'No')
        paid_cell = ws3.cell(row=row_idx, column=5, value=paid_val)
        if paid_val.startswith('Yes'):
            paid_cell.fill = YES_FILL
        else:
            paid_cell.fill = NO_FILL

    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 45
    ws3.column_dimensions['D'].width = 15
    ws3.column_dimensions['E'].width = 20

    # --- Foglio 4: Anomalies ---
    ws4 = wb.create_sheet('4. Anomalies')
    write_header(ws4, ['CustomerID', 'Type', 'EMEA Amount', 'Generated Amount', 'Difference', 'Detail'])
    for row_idx, an in enumerate(anomalies, 2):
        ws4.cell(row_idx, 1, an['CustomerID']); ws4.cell(row_idx, 2, an['Type'])
        ws4.cell(row_idx, 3, an['EMEA Amount']); ws4.cell(row_idx, 4, an['Generated Amount'])
        ws4.cell(row_idx, 5, an['Difference']); ws4.cell(row_idx, 6, an['Detail'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
